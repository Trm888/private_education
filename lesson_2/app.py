import asyncio
import os
import pathlib

from eth_account import Account
from openpyxl import Workbook
from web3 import Web3

from sdk.client import Client
from sdk.data.models import Networks


# from private_data import private_key1, private_key2, private_key3, proxy
#
#
# async def main():
#     client = Client(private_key=private_key1, network=Networks.Optimism, proxy=proxy)
#     print(await client.wallet.balance(token_address='0xaf88d065e77c8cc2239327c5edb3a432268e5831'))
#     balance = await client.wallet.balance()
#     balance = await client.wallet.balance()
#     balance = await client.wallet.balance()

async def create_check_wallet():
    client = Client(network=Networks.Ethereum)
    balance = await client.wallet.balance()
    print(f'address: {client.account.address}, balance: {balance}')


async def create_wallet_my(qty):
    web3 = Web3(Web3.HTTPProvider('https://rpc.ankr.com/eth/'))
    wb = Workbook()
    ws = wb.active
    ws.title = "Wallets"
    ws["A1"] = "Address"
    ws["B1"] = "Private Key"
    for _ in range(qty):
        private_key = Web3.to_hex(os.urandom(32))
        account = Account.from_key(private_key)
        checksum_address = Web3.to_checksum_address(account.address)
        balance = web3.eth.get_balance(checksum_address)
        print(f"balance of {account.address} : {balance} Wei")

        ws.append([account.address, private_key])
    filepath = pathlib.PurePath.joinpath(pathlib.Path.cwd(), 'wallets.xlsx')
    wb.save(filepath)


async def main(qty):
    tasks = []
    for _ in range(qty):
        tasks.append(asyncio.create_task(create_check_wallet()))
    await asyncio.wait(tasks)


if __name__ == '__main__':
    loop = asyncio.new_event_loop()
    loop.run_until_complete(main(10))
